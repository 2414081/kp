import json
import asyncio
import os
import argparse
import pandas as pd
import re
from pathlib import Path
import sys
import logging
from logging.handlers import RotatingFileHandler

# Ensure project root is on sys.path so 'configs' and other top-level packages can be imported
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.append(str(PROJECT_ROOT))

from configs.llm_config import chat_completions
from prompts.metric_prompts.prompt_1 import (
    groundtruth_accuracy_prompt,
    completeness_evaluation_prompt,
    context_precision_prompt,
    context_recall_prompt,
    hallucination_consistency_prompt 
)
from prompts.metric_prompts.prompt_2 import hallucination_cove_prompt

# Create logs directory (and subdirectory) if they don't exist
os.makedirs("logs/metric_evaluation_logs", exist_ok=True)

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
logger.propagate = False  # avoid duplicate logs if root has handlers

if not logger.handlers:
    file_handler = RotatingFileHandler(
        filename=r"logs/metric_evaluation_logs/app.log",
        mode="a",
        maxBytes=5 * 1024 * 1024,  # 5 MB
        backupCount=3,
        encoding="utf-8"
    )
    formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler(stream=sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

logger.info("metric_evaluation logger initialized")

# Load prompts from Excel file
def load_prompts_from_excel(file_path):
    """
    Load prompts from an Excel file and validate the structure.
    """
    try:
        # Read the Excel file into a DataFrame
        df = pd.read_excel(file_path)

        # Ensure required columns exist
        required_columns = ["question", "ground_truth_answer", "retrieval_context_chat"]
        if not all(col in df.columns for col in required_columns):
            print(f"Error: Missing required columns in the Excel file. Required columns: {required_columns}")
            return []

        # Convert DataFrame to a list of dictionaries
        return df.to_dict(orient="records")
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return []

# Helper: Call LLM and get response
async def get_llm_response(prompt):
    """
    Send a prompt to the LLM and return the response.
    """
    messages = [{"role": "user", "content": prompt}]
    try:
        result = await chat_completions(messages)
        response = result.get('final_result', '')
        # Debug: always show raw
        # print(f"[DEBUG] Raw LLM text: {response}")

        # Normalize fenced code blocks
        if isinstance(response, str) and response.startswith("```"):
            # Remove leading and trailing fences
            response = response.strip().strip("`")
            # Remove leading language tag like json\n
            response = re.sub(r"^(json|python|txt)\s*", "", response, flags=re.IGNORECASE)

        # Try direct JSON load
        if isinstance(response, str):
            try:
                return json.loads(response)
            except json.JSONDecodeError:
                # Try to extract first JSON object/array substring
                match = re.search(r"(\{.*\}|\[.*\])", response, re.DOTALL)
                if match:
                    candidate = match.group(1)
                    try:
                        return json.loads(candidate)
                    except json.JSONDecodeError:
                        print(f"[WARN] Extracted candidate not valid JSON.")
                print(f"[WARN] LLM response not valid JSON; returning raw string instead of None.")
                return response  # Fallback: return raw string so caller can log it
        return response
    except Exception as e:
        print(f"Error while calling LLM: {e}")
        return None

# Evaluate prompts
async def evaluate_prompts(file_path, excel_output_folder, selected_prompts):
    """
    Evaluate prompts using the LLM and calculate metrics for the selected prompt functions.
    (Incremental append version)
    """
    dataset = load_prompts_from_excel(file_path)
    if not dataset:
        print("No records loaded.")
        return

    total_entries = len(dataset)
    processed_entries = 0
    prompt_success_counts = {p: 0 for p in selected_prompts}

    print(f"Total records loaded: {total_entries}")

    output_file = f"{excel_output_folder}/metric_evaluation_results_2.xlsx"
    os.makedirs(excel_output_folder, exist_ok=True)

    for entry_idx, entry in enumerate(dataset, start=1):
        question = entry.get("question")
        ground_truth_answer = entry.get("ground_truth_answer")
        actual_chat_response_1 = entry.get("actual_chat_response_1")
        actual_chat_response_2 = entry.get("actual_chat_response_2")
        actual_chat_response_3 = entry.get("actual_chat_response_3")
        chat_generated_questions = entry.get("chat_generated_questions")
        retrieval_context_chat = entry.get("retrieval_context_chat")
        
        if not question or not ground_truth_answer or not retrieval_context_chat:
            print(f"Record {entry_idx}: skipped (missing required fields).")
            continue

        processed_entries += 1
        evaluations = {}

        if "groundtruth_accuracy_prompt" in selected_prompts:
            prompt = groundtruth_accuracy_prompt(
                ground_truth_answer=ground_truth_answer, actual_response=actual_chat_response_1)
            evaluations["groundtruth_accuracy_prompt"] = await get_llm_response(prompt)
        if "completeness_evaluation_prompt" in selected_prompts:
            prompt = completeness_evaluation_prompt(
                ground_truth_answer=ground_truth_answer, actual_response=actual_chat_response_1, question=question)
            evaluations["completeness_evaluation_prompt"] = await get_llm_response(prompt)
        if "context_precision_prompt" in selected_prompts:
            prompt = context_precision_prompt(retrieval_context_chat, question)
            evaluations["context_precision_prompt"] = await get_llm_response(prompt)
        if "context_recall_prompt" in selected_prompts:
            prompt = context_recall_prompt(ground_truth_answer, retrieval_context_chat)
            evaluations["context_recall_prompt"] = await get_llm_response(prompt)
        if "hallucination_cove_prompt" in selected_prompts:
            required = {
                "ground_truth_answer": ground_truth_answer,
                "actual_chat_response_1": actual_chat_response_1,
                "chat_generated_questions": chat_generated_questions,
            }
            missing = [k for k, v in required.items() if not v]
            if not missing:
                prompt = hallucination_cove_prompt(
                    actual_chat_response_1=actual_chat_response_1,
                    verification_questions=chat_generated_questions,
                    ground_truth_answer=ground_truth_answer
                )
                evaluations["hallucination_cove_prompt"] = await get_llm_response(prompt)
            else:
                print(f"Record {entry_idx} [hallucination_cove_prompt] skipped (missing fields: {', '.join(missing)}).")
        if "hallucination_consistency_prompt" in selected_prompts:
            prompt = hallucination_consistency_prompt(
                actual_answer_1=actual_chat_response_1,
                actual_answer_2=actual_chat_response_2,
                actual_answer_3=actual_chat_response_3
            )
            evaluations["hallucination_consistency_prompt"] = await get_llm_response(prompt)

        # Process & append immediately
        for prompt_name, eval_prompt in evaluations.items():
            print(f"Record {entry_idx} [{prompt_name}] LLM raw response: {eval_prompt}")
            if eval_prompt is None:
                print(f"Record {entry_idx} [{prompt_name}] skipped (LLM returned None).")
                continue
            result = process_response(
                prompt_name,
                eval_prompt,
                file_path,
                question,
                ground_truth_answer,
                retrieval_context_chat,
                actual_chat_response_1,
                entry
            )
            append_results_to_excel_single(output_file, prompt_name, result)
            prompt_success_counts[prompt_name] += 1

    print("----- Evaluation Summary -----")
    print(f"Total records in dataset: {total_entries}")
    print(f"Records processed (with required fields): {processed_entries}")
    for p in selected_prompts:
        print(f"{p}: successful evaluations = {prompt_success_counts.get(p, 0)}")
    print(f"Incremental results saved to: {output_file}")

def append_results_to_excel_single(output_file, prompt_name, result):
    """
    Append (or update) one processed prompt result to workbook.
    - One sheet per prompt (detail rows always appended)
    - Main sheet: single consolidated row per question; metrics columns filled incrementally
    """
    from openpyxl import Workbook, load_workbook

    items = result.get("response_items")
    base = {k: v for k, v in result.items() if k != "response_items"}
    expanded_rows = []
    if isinstance(items, list) and items:
        for item in items:
            row = base.copy()
            row.update(item)
            expanded_rows.append(row)
    else:
        expanded_rows.append(base)

    score_key_map = {
        "groundtruth_accuracy_prompt": "Groundtruth Accuracy",
        "completeness_evaluation_prompt": "Completeness",
        "context_precision_prompt": "Context Precision",
        "context_recall_prompt": "Context Recall",
        "hallucination_cove_prompt": "Hallucination CoVe",
        "hallucination_consistency_prompt": "Hallucination Consistency Score"
    }
    score_header = score_key_map.get(prompt_name)

    if not os.path.exists(output_file):
        wb = Workbook()
        main_ws = wb.active
        main_ws.title = "Main"
        main_headers = [
            "topic","sub_topic","section","question_type","question",
            "ground_truth_answer",
            "actual_chat_response_1",
            "actual_chat_response_2",
            "actual_chat_response_3",
            "chat_generated_questions",
            "retrieval_context_chat",
            "Groundtruth Accuracy","Completeness","Context Precision",
            "Context Recall","Hallucination CoVe","Hallucination Consistency Score"
        ]
        main_ws.append(main_headers)
        wb.create_sheet(prompt_name[:31])
        wb.save(output_file)

    wb = load_workbook(output_file)

    # ---- Prompt detail sheet (fixed column alignment) ----
    sheet_name = prompt_name[:31]
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    prompt_ws = wb[sheet_name]

    existing_headers = []
    if prompt_ws.max_row >= 1:
        existing_headers = [
            prompt_ws.cell(row=1, column=c).value
            for c in range(1, prompt_ws.max_column + 1)
            if prompt_ws.cell(row=1, column=c).value
        ]

    # Stable base header order
    base_header_order = [
        "topic","sub_topic","section","question_type","question",
        "ground_truth_answer",
        "actual_chat_response_1",
        "actual_chat_response_2",
        "actual_chat_response_3",
        "chat_generated_questions",
        "retrieval_context_chat",
        "Raw LLM Response",
        # Metric-specific possible fields (pre-listed for stability)
        "Fully Matched","Partially Matched","Total Statements","Groundtruth Accuracy",
        "Fully Covered","Partially Covered","Total Expected Elements","Completeness",
        "Relevant Fragments","Partially Relevant Fragments","Total Fragments","Context Precision",
        "Contained Facts","Partially Contained Facts","Total Factual Statements","Context Recall",
        "Total Questions","Correct Answers","Hallucination CoVe",
        "Consistent Statements","Hallucination Consistency Score",
        # Generic aggregation fields used by process_response
        "response_items"
    ]

    # Collect appearance-ordered new keys from expanded rows
    appearance_order = []
    for r in expanded_rows:
        for k in r.keys():
            if k not in appearance_order:
                appearance_order.append(k)

    if not existing_headers:
        # First time: start with base order that actually appears, then append remaining appearing keys
        full_headers = [h for h in base_header_order if h in appearance_order]
        for k in appearance_order:
            if k not in full_headers:
                full_headers.append(k)
    else:
        # Preserve existing header order; append new keys in first-seen order
        full_headers = existing_headers[:]
        for k in appearance_order:
            if k not in full_headers:
                full_headers.append(k)

    # Write / extend header row only if changed
    if not existing_headers or len(full_headers) != len(existing_headers):
        for col_idx, h in enumerate(full_headers, 1):
            prompt_ws.cell(row=1, column=col_idx).value = h

    # Append detail rows aligned with full_headers
    start_detail_row = prompt_ws.max_row + 1
    for r in expanded_rows:
        row_idx = prompt_ws.max_row + 1
        for col_idx, h in enumerate(full_headers, 1):
            prompt_ws.cell(row=row_idx, column=col_idx).value = r.get(h)

    # ---- Main sheet (update / consolidate) ----
    main_ws = wb["Main"]
    main_headers = [main_ws.cell(row=1, column=c).value for c in range(1, main_ws.max_column + 1)]

    if score_header and score_header not in main_headers:
        main_headers.append(score_header)
        main_ws.cell(row=1, column=len(main_headers)).value = score_header

    main_row_dict = {
        "topic": result.get("topic"),
        "sub_topic": result.get("sub_topic"),
        "section": result.get("section"),
        "question_type": result.get("question_type"),
        "question": result.get("question"),
        "ground_truth_answer": result.get("ground_truth_answer"),
        "actual_chat_response_1": result.get("actual_chat_response_1"),
        "actual_chat_response_2": result.get("actual_chat_response_2"),
        "actual_chat_response_3": result.get("actual_chat_response_3"),
        "chat_generated_questions": result.get("chat_generated_questions"),
        "retrieval_context_chat": result.get("retrieval_context_chat"),
    }
    if score_header:
        main_row_dict[score_header] = result.get(score_header, "N/A")

    # Format metric scores to two decimals
    metrics_to_format = [
        "Groundtruth Accuracy","Completeness","Context Precision",
        "Context Recall","Hallucination CoVe","Hallucination Consistency Score"
    ]
    for m in metrics_to_format:
        v = main_row_dict.get(m)
        if isinstance(v, (int, float)):
            main_row_dict[m] = round(v, 2)

    question_col_index = main_headers.index("question") + 1
    existing_row_index = None
    target_question = main_row_dict.get("question")

    for r in range(2, main_ws.max_row + 1):
        if main_ws.cell(row=r, column=question_col_index).value == target_question:
            existing_row_index = r
            break

    if existing_row_index:
        for h_idx, header in enumerate(main_headers, start=1):
            new_val = main_row_dict.get(header)
            if new_val is not None:
                cell = main_ws.cell(row=existing_row_index, column=h_idx)
                cell.value = new_val
                if header in metrics_to_format and isinstance(new_val, (int, float)):
                    cell.number_format = "0.00"
        if score_header:
            score_col_index = main_headers.index(score_header) + 1
            cell = main_ws.cell(row=existing_row_index, column=score_col_index)
            cell.hyperlink = f"#{sheet_name}!A{start_detail_row}"
            cell.style = "Hyperlink"
    else:
        row_values = []
        for h in main_headers:
            val = main_row_dict.get(h)
            row_values.append(val)
        main_ws.append(row_values)
        new_row_index = main_ws.max_row
        for h_idx, header in enumerate(main_headers, start=1):
            cell = main_ws.cell(row=new_row_index, column=h_idx)
            if header in metrics_to_format and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
        if score_header:
            score_col_index = main_headers.index(score_header) + 1
            cell = main_ws.cell(row=new_row_index, column=score_col_index)
            cell.hyperlink = f"#{sheet_name}!A{start_detail_row}"
            cell.style = "Hyperlink"

    wb.save(output_file)
    print(f"[APPEND] {prompt_name} saved; detail rows start at {start_detail_row} (main row {'updated' if existing_row_index else 'added'}).")

# Process the response for each prompt
def process_response(prompt_name, response, file_path, question, ground_truth_answer, retrieval_context_chat, actual_chat_response_1, input_row):
    result = input_row.copy()
    result["Raw LLM Response"] = response if isinstance(response, str) else json.dumps(response, indent=2)

    if not isinstance(response, list):
        return result

    if prompt_name == "groundtruth_accuracy_prompt":
        fully_matched = sum(1 for item in response if item.get("Match Status", "").lower() == "matched")
        partially_matched = sum(1 for item in response if item.get("Match Status", "").lower() == "partially matched")
        total_statements = len(response) or 1
        groundtruth_accuracy = (fully_matched + 0.5 * partially_matched) / total_statements
        result.update({
            "Fully Matched": fully_matched,
            "Partially Matched": partially_matched,
            "Total Statements": total_statements,
            "Groundtruth Accuracy": groundtruth_accuracy
        })
    elif prompt_name == "completeness_evaluation_prompt":
        fully_covered = sum(1 for item in response if item.get("Coverage Label", "").strip().lower() == "covered")
        partially_covered = sum(1 for item in response if item.get("Coverage Label", "").strip().lower() == "partially covered")
        total_expected_elements = len(response) or 1
        completeness = (fully_covered + 0.5 * partially_covered) / total_expected_elements
        result.update({
            "Fully Covered": fully_covered,
            "Partially Covered": partially_covered,
            "Total Expected Elements": total_expected_elements,
            "Completeness": completeness
        })
    elif prompt_name == "context_precision_prompt":
        relevant_fragments = sum(1 for item in response if item.get("Relevance Label", "").strip().lower() == "relevant")
        partially_relevant_fragments = sum(1 for item in response if item.get("Relevance Label", "").strip().lower() == "partially relevant")
        total_fragments = len(response) or 1
        context_precision = (relevant_fragments + 0.5 * partially_relevant_fragments) / total_fragments
        result.update({
            "Relevant Fragments": relevant_fragments,
            "Partially Relevant Fragments": partially_relevant_fragments,
            "Total Fragments": total_fragments,
            "Context Precision": context_precision
        })
    elif prompt_name == "context_recall_prompt":
        contained_facts = sum(1 for item in response if item.get("Label", "").strip().lower() == "contained")
        partially_contained_facts = sum(1 for item in response if item.get("Label", "").strip().lower() == "partially contained")
        total_factual_statements = len(response) or 1
        context_recall = (contained_facts + 0.5 * partially_contained_facts) / total_factual_statements
        result.update({
            "Contained Facts": contained_facts,
            "Partially Contained Facts": partially_contained_facts,
            "Total Factual Statements": total_factual_statements,
            "Context Recall": context_recall
        })
    elif prompt_name == "hallucination_cove_prompt":
        correct_answers = sum(1 for item in response if item.get("derived_expected_answer") == item.get("derived_actual_answer"))
        total_questions = len(response) or 1
        raw_score = correct_answers / total_questions  # original (0 = worst, 1 = best)
        hallucination_cove = 1 - raw_score              # inverted (0 = best, 1 = worst)
        result.update({
            "Total Questions": total_questions,
            "Correct Answers": correct_answers,
            "Hallucination CoVe": hallucination_cove
        })
    elif prompt_name == "hallucination_consistency_prompt":
        consistent_statements = sum(1 for item in response if item.get("Consistency Label", "").lower() == "consistent")
        total_factual_statements = len(response) or 1
        raw_score = consistent_statements / total_factual_statements  # original
        hallucination_consistency_score = 1 - raw_score               # inverted
        result.update({
            "Consistent Statements": consistent_statements,
            "Total Factual Statements": total_factual_statements,
            "Hallucination Consistency Score": hallucination_consistency_score
        })

    # Store raw list for later sheet expansion
    result["response_items"] = response
    return result

# Main entry point
if __name__ == "__main__":
    # Parse command-line arguments
    parser = argparse.ArgumentParser(description="Evaluate prompts fr om Excel files.")
    parser.add_argument(
        "--prompts",
        nargs="+",
        choices=[
            "groundtruth_accuracy_prompt",
            "completeness_evaluation_prompt",
            "context_precision_prompt",
            "context_recall_prompt",
            "hallucination_consistency_prompt",
            "hallucination_cove_prompt",
            "all"
        ],
        default=["all"],
        help="Specify which prompts to evaluate. Use 'all' to evaluate all prompts."
    )
    args = parser.parse_args()

    # Path to the Excel file
    file_path = "outputs/data_preparation_output_samples/combined_updated_questions_5g_r.xlsx"
    excel_output_folder = r"P:\test_new_2\results"
    os.makedirs(excel_output_folder, exist_ok=True)

    # Filter prompts based on user input
    selected_prompts = args.prompts
    if "all" in selected_prompts:
        selected_prompts = [
            "groundtruth_accuracy_prompt",
            "completeness_evaluation_prompt",
            "context_precision_prompt",
            "context_recall_prompt",
            "hallucination_consistency_prompt",
            "hallucination_cove_prompt"
        ]

    # Run the evaluation
    asyncio.run(evaluate_prompts(file_path, excel_output_folder, selected_prompts))
