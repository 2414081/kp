import pandas as pd
from pathlib import Path
import re
from openpyxl import load_workbook

INPUT_PATH = r"outputs/metrics_output_samples/metric_evaluation_results_2_genetics.xlsx"
SHEET_NAME = "Main"  # change if needed
PARENT_COL = "Parent_Index"
PARENT_PREFIX = "PTC_"
CHILD_COL = "Child_Index"  # new
IS_PARENT_COL = "Is_Parent"  # optional flag

# Thresholds for pass/fail
PASS_FAIL_THRESHOLDS = {
    "groundtruth_accuracy_prompt": 0.75,
    "completeness_evaluation_prompt": 0.70,
    "context_precision_prompt": 0.75,
    "context_recall_prompt": 0.75,
    "hallucination_cove_prompt": 0.20,
    "hallucination_consistency_prompt": 0.20,
}

# Metrics where higher is better (>= threshold is pass)
HIGHER_IS_BETTER = [
    "groundtruth_accuracy_prompt",
    "completeness_evaluation_prompt",
    "context_precision_prompt",
    "context_recall_prompt",
]

# Metrics where lower is better (<= threshold is pass)
LOWER_IS_BETTER = [
    "hallucination_cove_prompt",
    "hallucination_consistency_prompt",
]

METRIC_MAP = {
    "Groundtruth Accuracy": "groundtruth_accuracy_prompt",
    "Completeness": "completeness_evaluation_prompt",
    "Context Precision": "context_precision_prompt",
    "Context Recall": "context_recall_prompt",  # corrected key casing
    "Hallucination Consistency Score": "hallucination_consistency_prompt",  # corrected mapping
    "Hallucination CoVe": "hallucination_cove_prompt",  # corrected mapping
}

def assign_parent_groups(df: pd.DataFrame, type_col: str = "question_type") -> pd.DataFrame:
    if type_col not in df.columns:
        raise KeyError(f"Column '{type_col}' not found in the sheet.")

    # Parent groups: start at each 'Base'
    parent_id = df[type_col].eq("Base").cumsum()
    parent_prefix = "PT_" + parent_id.astype(int).astype(str).str.zfill(2)

    # Assign Parent_Index based on type
    df[PARENT_COL] = parent_prefix + "_" + df.groupby(parent_id).cumcount().map(
        lambda x: "B" if x == 0 else f"C{x:02}"
    )

    return df

def normalize_metric_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Rename columns based on known headers in the Excel
    rename_map = {src: dst for src, dst in METRIC_MAP.items() if src in df.columns}
    if rename_map:
        df = df.rename(columns=rename_map)
    return df

def apply_metrics_status(df: pd.DataFrame) -> pd.DataFrame:
    # Normalize metric column names first
    df = normalize_metric_columns(df)

    # Ensure all metric columns exist
    missing = [m for m in PASS_FAIL_THRESHOLDS.keys() if m not in df.columns]
    if missing:
        raise KeyError(f"Missing metric columns in sheet: {missing}")

    # Compute per-metric pass/fail booleans without adding columns
    checks = []
    for m in HIGHER_IS_BETTER:
        thr = PASS_FAIL_THRESHOLDS[m]
        checks.append(df[m].ge(thr))
    for m in LOWER_IS_BETTER:
        thr = PASS_FAIL_THRESHOLDS[m]
        checks.append(df[m].le(thr))

    # Overall status: pass only if all metrics pass
    df["metrics_status"] = pd.concat(checks, axis=1).all(axis=1).map({True: "Pass", False: "Fail"})
    return df

def add_robustness_score(df: pd.DataFrame) -> pd.DataFrame:
    # Map Pass/Fail to 1/0 for scoring
    status_numeric = df["metrics_status"].map({"Pass": 1, "Fail": 0}).astype(float)

    # Weights for base and child scores
    base_weight = 0.0
    child_weight = 1.0

    # Calculate scores for each Parent_Index group
    group_scores = {}
    for parent, group in df.groupby(df[PARENT_COL].str.extract(r"(PT_\d+)_")[0]):
        print(f"Processing parent group: {parent}")  # Debug

        base_rows = group[group[PARENT_COL].str.endswith("_B")]
        child_rows = group[~group[PARENT_COL].str.endswith("_B")]

        base_score = status_numeric.loc[base_rows.index].mean() * base_weight
        child_scores = (
            status_numeric.loc[child_rows.index].mean() * child_weight
            if not child_rows.empty
            else 0
        )

        base_score = base_score if not pd.isna(base_score) else 0
        child_scores = child_scores if not pd.isna(child_scores) else 0

        group_scores[parent] = round(base_score + child_scores, 2)

        # Debugging scores
        print(f"Parent: {parent}")
        print(f"Base score: {base_score}")
        print(f"Child scores: {child_scores}")
        print(f"Total group score: {group_scores[parent]}")

    # Skip normalization
    # max_score = max(group_scores.values()) if group_scores else 1
    # max_score = max(max_score, 1e-6)  # Avoid division by zero or very small values
    # group_scores = {k: round(v / max_score, 2) for k, v in group_scores.items()}

    # Assign robustness scores only to base rows
    def map_robustness_score(x):
        match = re.match(r"(PT_\d+)_", x)
        if match:
            parent_group = match.group(1)
            score = group_scores.get(parent_group, 0)
            print(f"Mapping {x} -> Parent group: {parent_group}, Score: {score}")  # Debug
            return score if x.endswith("_B") else float("nan")
        else:
            print(f"Mapping {x} -> No match found, Score: 0")  # Debug
            return float("nan")

    df["robustness_score"] = df[PARENT_COL].map(map_robustness_score)

    print(df[[PARENT_COL, "robustness_score"]])  # Debug

    return df

def add_group_pass_status(df: pd.DataFrame) -> pd.DataFrame:
    # Ensure metrics_status exists
    if "metrics_status" not in df.columns:
        raise KeyError("Column 'metrics_status' not found. Run apply_metrics_status first.")

    # Compute group-wise all-pass boolean keyed by parent group "PT_##"
    parent_groups = df[PARENT_COL].str.extract(r"(PT_\d+)_")[0]
    all_pass_by_group = (
        df.assign(parent_group=parent_groups)
          .groupby("parent_group")["metrics_status"]
          .apply(lambda s: (s == "Pass").all())
    )

    # Map to base rows only; children get NaN
    def map_group_pass(x: str) -> float | str:
        m = re.match(r"(PT_\d+)_", x)
        if not m:
            return float("nan")
        grp = m.group(1)
        val = "Pass" if all_pass_by_group.get(grp, False) else "Fail"
        return val if x.endswith("_B") else float("nan")

    df["group_pass_status"] = df[PARENT_COL].map(map_group_pass)
    return df

def get_indexed_output_path(input_path: str) -> Path:
    p = Path(input_path)
    out_dir = p.parent
    stem = p.stem
    suffix = p.suffix or ".xlsx"

    # Base output name
    base = out_dir / f"{stem}_robust{suffix}"
    if not base.exists():
        return base

    # If exists, increment index
    i = 1
    while True:
        candidate = out_dir / f"{stem}_robust_{i}{suffix}"
        if not candidate.exists():
            return candidate
        i += 1

def write_new_columns_preserve_sheet(input_path: str, sheet_name: str, df: pd.DataFrame, output_path: Path) -> None:
    """
    Preserve the original sheet (including hyperlinks) and append/update only the new
    computed columns for the given sheet_name using openpyxl.
    """
    wb = load_workbook(input_path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in workbook.")
    ws = wb[sheet_name]

    # Build a mapping of existing headers to column indices
    headers = {}
    header_row = 1
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if isinstance(val, str):
            headers[val] = col

    # Ensure the DataFrame rows align with the worksheet rows:
    # Assumes that the DataFrame was parsed from the same sheet without row filtering,
    # so row i in df corresponds to Excel row i+2 (header at row 1).
    start_row = 2

    # Columns to add/update (only our computed ones)
    new_cols = []
    for col in ["metrics_status", "robustness_score", "group_pass_status"]:
        if col in df.columns:
            new_cols.append(col)

    # Assign columns for any new headers at the end
    next_col = max_col + 1
    for col_name in new_cols:
        if col_name not in headers:
            headers[col_name] = next_col
            ws.cell(row=header_row, column=next_col, value=col_name)
            next_col += 1

    # Write values for each new column, only for base rows when required
    for idx, row in df.iterrows():
        excel_row = start_row + idx

        # metrics_status: write for all rows
        if "metrics_status" in new_cols:
            c = ws.cell(row=excel_row, column=headers["metrics_status"])
            c.value = row.get("metrics_status")

        # robustness_score: only base rows have a score, children NaN -> leave blank
        if "robustness_score" in new_cols:
            val = row.get("robustness_score")
            c = ws.cell(row=excel_row, column=headers["robustness_score"])
            c.value = None if pd.isna(val) else float(val)

        # group_pass_status: only base rows, children blank
        if "group_pass_status" in new_cols:
            val = row.get("group_pass_status")
            c = ws.cell(row=excel_row, column=headers["group_pass_status"])
            # If NaN, leave blank to avoid overwriting anything
            c.value = None if (isinstance(val, float) and pd.isna(val)) else val

    # Save to output path, preserving original hyperlinks
    wb.save(output_path)

def main():
    # Parse with pandas for computation
    sheets_df = pd.read_excel(INPUT_PATH, sheet_name=None)

    if SHEET_NAME not in sheets_df:
        # Show available sheets parsed by pandas
        raise KeyError(f"Sheet '{SHEET_NAME}' not found. Available sheets: {list(sheets_df.keys())}")

    df = sheets_df[SHEET_NAME]

    df = assign_parent_groups(df, type_col="question_type")
    df = apply_metrics_status(df)
    df = add_robustness_score(df)
    df = add_group_pass_status(df)  # base-only Pass/Fail for group

    output_path = get_indexed_output_path(INPUT_PATH)

    # Preserve hyperlinks by writing only new columns via openpyxl
    write_new_columns_preserve_sheet(INPUT_PATH, SHEET_NAME, df, output_path)

    # For other sheets, copy as-is to the new workbook
    # If you need to copy all sheets, you can extend write_new_columns_preserve_sheet accordingly.

    print(f"Done. Output file: {output_path}")


if __name__ == "__main__":
    main()