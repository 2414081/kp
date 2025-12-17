from pathlib import Path

import sys
# Ensure project root is on sys.path so 'configs' and other top-level packages can be imported
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.append(str(PROJECT_ROOT))

from configs.llm_config import chat_completions
import json
import asyncio
import re
import logging
from logging.handlers import RotatingFileHandler
from html.parser import HTMLParser
from openpyxl import Workbook, load_workbook
from prompts.test_data_generation_prompts.prompt import section_prompt, combined_prompt

# --- Logging setup ---
LOG_DIR = PROJECT_ROOT / "logs" / "test_data_generation_logs"  # as requested name; created if not exists
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = LOG_DIR / "app.log"

logger = logging.getLogger("test_data_generation")
logger.setLevel(logging.INFO)

# File handler with rotation (5 MB per file, keep 3 backups)
_file_handler = RotatingFileHandler(LOG_FILE, maxBytes=5_000_000, backupCount=3, encoding="utf-8")
_file_handler.setLevel(logging.INFO)
_file_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))

# Console handler
_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setLevel(logging.INFO)
_console_handler.setFormatter(logging.Formatter("%(levelname)s | %(message)s"))

# Attach handlers once
if not logger.handlers:
    logger.addHandler(_file_handler)
    logger.addHandler(_console_handler)

def log_info(msg: str):
    logger.info(msg)

def log_error(msg: str, exc: Exception | None = None):
    if exc:
        logger.error(f"{msg} | Exception: {exc}", exc_info=True)
    else:
        logger.error(msg)

CAREGUIDES_INPUT = Path(r"inputs/careguides_input_samples")
OUTPUT_DIR = Path(r"outputs/test_data_generation_output_samples")

# COMBINED_OUTPUT_DIR no longer used as a top-level; keep for backward compatibility if referenced elsewhere
COMBINED_OUTPUT_DIR = OUTPUT_DIR / "combined_output"
INDIVIDUAL_OUTPUT_DIR = OUTPUT_DIR / "individual_output"
INDIVIDUAL_XLSX = OUTPUT_DIR / "individual_output.xlsx"
COMBINED_XLSX = OUTPUT_DIR / "combined_output.xlsx"

async def get_llm_response(prompt: str):
    messages = [{"role": "user", "content": prompt}]
    try:
        result = await chat_completions(messages)
        return result.get("final_result", "")
    except Exception as e:
        log_error("LLM call failed in get_llm_response", e)
        return ""

class HTMLTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows = []               # list[list[dict]] where dict = {"tag": "th"/"td", "attrs": dict, "text": str, "html": str}
        self._in_tr = False
        self._in_cell = False
        self._cell_tag = None
        self._cell_attrs = {}
        self._cell_text_parts = []
        self._cell_html_parts = []
        self._current_row = []
        self._tag_stack = []

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._in_tr = True
            self._current_row = []
        elif tag in ("th", "td"):
            self._in_cell = True
            self._cell_tag = tag
            self._cell_attrs = dict(attrs)
            self._cell_text_parts = []
            self._cell_html_parts = [f"<{tag}" + "".join([f' {k}="{v}"' for k, v in attrs]) + ">"]
        elif self._in_cell:
            # Keep inner HTML roughly
            self._cell_html_parts.append(f"<{tag}>")
        self._tag_stack.append(tag)

    def handle_endtag(self, tag):
        # close inner tags
        if self._in_cell and tag not in ("tr",):
            self._cell_html_parts.append(f"</{tag}>")
        # closing cell
        if tag in ("th", "td") and self._in_cell and self._cell_tag == tag:
            text = " ".join(part.strip() for part in self._cell_text_parts if part.strip())
            html = "".join(self._cell_html_parts) + f"</{tag}>"
            self._current_row.append({"tag": tag, "attrs": self._cell_attrs, "text": text, "html": html})
            self._in_cell = False
            self._cell_tag = None
            self._cell_attrs = {}
            self._cell_text_parts = []
            self._cell_html_parts = []
        elif tag == "tr" and self._in_tr:
            self.rows.append(self._current_row)
            self._current_row = []
            self._in_tr = False
        # pop stack
        if self._tag_stack and self._tag_stack[-1] == tag:
            self._tag_stack.pop()

    def handle_data(self, data):
        if self._in_cell:
            self._cell_text_parts.append(data)
            self._cell_html_parts.append(data)

def extract_subsections_from_cell_html(cell_html: str, cell_text_fallback: str):
    """Extract subsections from a table cell based on <strong> tags, using simple HTML parsing"""
    subsections = []

    # Quick skip for short or N/A
    if not cell_html or cell_text_fallback.strip().upper() == "N/A" or len(cell_text_fallback.strip()) < 10:
        return subsections

    # naive find all <strong>...</strong>
    strong_spans = []
    for match in re.finditer(r"(<strong[^>]*>)(.*?)(</strong>)", cell_html, flags=re.IGNORECASE | re.DOTALL):
        start_tag, inner, end_tag = match.groups()
        strong_spans.append({"start": match.start(), "end": match.end(), "title": re.sub(r"\s+", " ", inner).strip(), "html": match.group(0)})

    if not strong_spans:
        # no strong, use entire cell text
        text = normalize_text(strip_tags(cell_html))
        if text and len(text) > 10:
            subsections.append({"section_title": "General", "section_text": text})
        return subsections

    # Split content between strong tags
    for idx, span in enumerate(strong_spans):
        title = span["title"]
        if not title:
            continue

        # content range: from end of current strong to start of next strong or end of cell
        content_start = span["end"]
        content_end = strong_spans[idx + 1]["start"] if idx + 1 < len(strong_spans) else len(cell_html)
        section_html = cell_html[content_start:content_end]

        # remove leading punctuation and normalize
        section_text = normalize_text(strip_tags(section_html))
        section_text = re.sub(r'^[:\-\s]+', '', section_text)
        section_text = re.sub(r'\s+', ' ', section_text)

        if section_text and len(section_text) > 10:
            subsections.append({"section_title": title, "section_text": section_text})

    return subsections

def strip_tags(html: str) -> str:
    # very simple tag stripper; keeps text content
    return re.sub(r"<[^>]+>", " ", html or "").strip()

def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())

ALLOWED_SECTIONS = {
    "background": "background",
    "contraindication": "contraindication",
    "contraindications": "contraindication",
    "evaluation": "evaluation",
    "lifestyle": "lifestyle",
    "management": "management",
    "patient communication": "patient_communication",
    "patient_communication": "patient_communication",
    "referral": "referral",
}

def extract_sections(html: str):
    """Extract only allowed sections; map headers and <strong> subsections to ALLOWED_SECTIONS keys."""
    parser = HTMLTableParser()
    parser.feed(html)
    rows = parser.rows

    allowed_keys = set(ALLOWED_SECTIONS.values())  # normalized targets
    sections = []

    # 1) Full-width headers (e.g., Patient Communication)
    i = 0
    while i < len(rows):
        row = rows[i]
        headers = [c for c in row if c["tag"] == "th"]
        if headers and len(headers) == 1 and (headers[0]["attrs"].get("colspan") == "2"):
            header_text = normalize_text(headers[0]["text"])
            norm_key = ALLOWED_SECTIONS.get(header_text.lower())
            if norm_key in allowed_keys and i + 1 < len(rows):
                content_row = rows[i + 1]
                cells = [c for c in content_row if c["tag"] == "td"]
                aggregated = " ".join(
                    t for t in [normalize_text(strip_tags(c["html"])) for c in cells]
                    if t and t.upper() != "N/A"
                ).strip()
                sections.append({
                    "section_name": norm_key,   # normalized allowed name only
                    "section_text": aggregated
                })
                i += 2
                continue
        i += 1

    # 2) Subsections within two-column cells mapped to allowed names (e.g., Background/Evaluation/Management/Referral)
    two_col_i = 0
    while two_col_i < len(rows):
        row = rows[two_col_i]
        headers = [c for c in row if c["tag"] == "th"]
        if headers and len(headers) == 2 and two_col_i + 1 < len(rows):
            content_row = rows[two_col_i + 1]
            cells = [c for c in content_row if c["tag"] == "td"]
            if len(cells) == 2:
                # Left cell subsections
                for sub in extract_subsections_from_cell_html(cells[0]["html"], cells[0]["text"]):
                    title_norm = ALLOWED_SECTIONS.get(sub["section_title"].lower())
                    if title_norm in allowed_keys:
                        sections.append({
                            "section_name": title_norm,
                            "section_text": sub.get("section_text", "")
                        })
                # Right cell subsections
                for sub in extract_subsections_from_cell_html(cells[1]["html"], cells[1]["text"]):
                    title_norm = ALLOWED_SECTIONS.get(sub["section_title"].lower())
                    if title_norm in allowed_keys:
                        sections.append({
                            "section_name": title_norm,
                            "section_text": sub.get("section_text", "")
                        })
                two_col_i += 2
                continue
        two_col_i += 1

    # 3) Deduplicate by normalized section_name; prefer non-empty text
    dedup = {}
    for s in sections:
        name = s["section_name"]
        text = s.get("section_text", "")
        if name not in dedup or (text and not dedup[name].get("section_text")):
            dedup[name] = {"section_name": name, "section_text": text}

    return list(dedup.values())

def _strip_code_fence(s: str) -> str:
    """Remove leading/trailing markdown code fences."""
    s = (s or "").strip()
    if s.startswith("```"):
        s = re.sub(r"^```[a-zA-Z]*\s*", "", s)
        s = re.sub(r"\s*```$", "", s)
    return s

def safe_json_loads(s: str):
    """
    Attempt to parse JSON from possibly noisy LLM output.
    - Strips code fences and BOM.
    - If direct parse fails, extracts the first {...} or [...] block.
    Returns (obj, err) where obj is parsed JSON or None, err is an exception or None.
    """
    if s is None:
        return None, ValueError("Response is None")

    s = _strip_code_fence(s)
    s = s.lstrip("\ufeff").strip()

    # Try direct parse first
    try:
        return json.loads(s), None
    except Exception as e:
        pass

    # Extract first JSON object or array via regex
    # This is a simple heuristic; it balances robustness vs. complexity.
    match = re.search(r"(\{.*\}|\[.*\])", s, flags=re.DOTALL)
    if match:
        candidate = match.group(1)
        try:
            return json.loads(candidate), None
        except Exception as e2:
            return None, e2

    return None, json.JSONDecodeError("No JSON object found", s, 0)

async def generate_questions_for_diagnosis(diagnosis_title: str, sections: list):
    """Generate questions for all allowed sections of a diagnosis"""
    log_info(f"Start diagnosis: {diagnosis_title} | sections={len(sections)}")
    diagnosis_sections = []
    for section in sections:
        section_name = section["section_name"]
        section_text = section["section_text"]
        log_info(f"Processing section: {section_name} | text_len={len(section_text)}")
        res = await generate_questions_for_section(diagnosis_title, section_name, section_text)
        items = res.get("items", [])
        if not isinstance(items, list) or not items or not isinstance(items[0], dict) or "basic_questions" not in items[0]:
            basic = res.get("basic_questions") or res.get("basic_question") or []
            if not isinstance(basic, list):
                basic = []
            items = [{
                "basic_questions": basic,   # prompt-only
                "validation_question": "",  # prompt-only; empty if missing
                "deepdive_question": ""     # prompt-only; empty if missing
            }]
        else:
            # Ensure keys exist but do not inject templates
            if "basic_questions" not in items[0] or not isinstance(items[0]["basic_questions"], list):
                items[0]["basic_questions"] = []
            if "validation_question" not in items[0] or not isinstance(items[0]["validation_question"], str):
                items[0]["validation_question"] = ""
            if "deepdive_question" not in items[0] or not isinstance(items[0]["deepdive_question"], str):
                items[0]["deepdive_question"] = ""
        diagnosis_sections.append({
            "section": section_name,
            "items": items
        })
    log_info(f"Completed diagnosis: {diagnosis_title}")
    return {"title": diagnosis_title, "sections": diagnosis_sections}

async def generate_questions_for_section(diagnosis_title: str, section_name: str, section_text: str):
    """Generate questions for a single section based on its aggregated text"""
    prompt_text = section_prompt(diagnosis_title, section_name, section_text)
    log_info(f"Generating questions for section: {section_name}")
    try:
        raw = await get_llm_response(prompt_text)
        response = (raw or "").strip()

        if not response:
            log_error(f"Empty LLM response for section {section_name}")
            raise json.JSONDecodeError("Empty response", "", 0)

        obj, err = safe_json_loads(response)
        if err:
            log_error(f"Error parsing LLM response for section {section_name}", err)
            log_info(f"Response snippet: {response[:200]}...")
            # Fallback skeleton
            obj = {"items": [{
                "basic_questions": [],
                "validation_question": "",
                "deepdive_question": ""
            }]}

        items = obj.get("items")

        # Normalize to expected structure without injecting templates
        if not isinstance(items, list) or not items or not isinstance(items[0], dict):
            obj["items"] = [{
                "basic_questions": [],
                "validation_question": "",
                "deepdive_question": ""
            }]
        else:
            entry = items[0]
            if "basic_questions" not in entry or not isinstance(entry.get("basic_questions"), list):
                entry["basic_questions"] = []
            if "validation_question" not in entry or not isinstance(entry.get("validation_question"), str):
                entry["validation_question"] = ""
            if "deepdive_question" not in entry or not isinstance(entry.get("deepdive_question"), str):
                entry["deepdive_question"] = ""

        # Retry once (prompt-only) if basic_questions empty
        if not obj["items"][0]["basic_questions"]:
            log_info("LLM returned empty basic_questions; retrying once...")
            raw_retry = await get_llm_response(prompt_text)
            response_retry = (raw_retry or "").strip()
            obj_retry, err_retry = safe_json_loads(response_retry)
            if not err_retry and isinstance(obj_retry, dict):
                items_retry = obj_retry.get("items")
                if isinstance(items_retry, list) and items_retry and isinstance(items_retry[0], dict):
                    if isinstance(items_retry[0].get("basic_questions"), list) and items_retry[0]["basic_questions"]:
                        obj["items"][0]["basic_questions"] = items_retry[0]["basic_questions"]
                    dd = items_retry[0].get("deepdive_question")
                    if isinstance(dd, str) and dd.strip():
                        obj["items"][0]["deepdive_question"] = dd.strip()
                    vq = items_retry[0].get("validation_question")
                    if isinstance(vq, str) and vq.strip():
                        obj["items"][0]["validation_question"] = vq.strip()
            else:
                log_error("Retry JSON parse failed in generate_questions_for_section", err_retry or None)

        return obj
    except Exception as e:
        log_error(f"Error parsing LLM response for section {section_name}", e)
        if 'response' in locals():
            log_info(f"Response snippet: {response[:200]}...")
        return {
            "diagnosis_title": diagnosis_title,
            "section": section_name,
            "items": [{
                "basic_questions": [],
                "validation_question": "",
                "deepdive_question": ""
            }]
        }

async def generate_combined_questions_for_sections(diagnosis_title: str, section_a: dict, section_b: dict):
    """
    Generate ONE combined clinical question for a pair of sections.
    Combines section texts when available; otherwise uses section names as fallback.
    """
    sec_a_name = section_a.get("section_name", "")
    sec_b_name = section_b.get("section_name", "")
    sec_a_text = section_a.get("section_text", "") or ""
    sec_b_text = section_b.get("section_text", "") or ""

    # Build combined context: prefer text; fallback to names when text is empty
    if sec_a_text.strip() or sec_b_text.strip():
        combined_text = " ".join(t for t in [sec_a_text.strip(), sec_b_text.strip()] if t).strip()
        combined_label = f"{sec_a_name} + {sec_b_name}"
    else:
        combined_text = ""
        combined_label = f"{sec_a_name} + {sec_b_name}"

    prompt_text = combined_prompt(diagnosis_title, combined_label, combined_text)
    log_info(f"Generating combined question: {diagnosis_title} | pair={combined_label}")
    try:
        raw = await get_llm_response(prompt_text)
        response = (raw or "").strip()
        if not response:
            log_error(f"Empty LLM response for combined {diagnosis_title} / {combined_label}")
            raise json.JSONDecodeError("Empty response", "", 0)

        obj_in, err = safe_json_loads(response)
        if err:
            log_error(f"Error parsing combined LLM response for {diagnosis_title} / {sec_a_name}+{sec_b_name}", err)
            log_info(f"Combined response snippet: {response[:200]}...")
            obj_in = {}

        # Normalize structure to ensure a single combined_question
        items = obj_in.get("items")
        cq = ""
        if isinstance(items, list) and items and isinstance(items[0], dict):
            cq = items[0].get("combined_question", "") or ""
        else:
            cq = obj_in.get("combined_question", "") or ""

        obj = {
            "diagnosis_title": diagnosis_title,
            "combined_section": combined_label,
            "items": [{
                "combined_question": cq.strip()
            }]
        }

        # Retry once if empty (prompt-only)
        if not obj["items"][0]["combined_question"]:
            log_info("Empty combined_question; retrying once...")
            raw_retry = await get_llm_response(prompt_text)
            response_retry = (raw_retry or "").strip()
            obj_retry, err_retry = safe_json_loads(response_retry)
            if not err_retry and isinstance(obj_retry, dict):
                items_retry = obj_retry.get("items")
                if isinstance(items_retry, list) and items_retry and isinstance(items_retry[0], dict):
                    cq_retry = items_retry[0].get("combined_question", "") or ""
                    if cq_retry.strip():
                        obj["items"][0]["combined_question"] = cq_retry.strip()
                else:
                    cq_retry = obj_retry.get("combined_question", "") or ""
                    if cq_retry.strip():
                        obj["items"][0]["combined_question"] = cq_retry.strip()
            else:
                log_error("Retry JSON parse failed in generate_combined_questions_for_sections", err_retry or None)

        return obj
    except Exception as e:
        log_error(f"Error parsing combined LLM response for {diagnosis_title} / {sec_a_name}+{sec_b_name}", e)
        if 'response' in locals():
            log_info(f"Combined response snippet: {response[:200]}...")
        return {
            "diagnosis_title": diagnosis_title,
            "combined_section": f"{sec_a_name} + {sec_b_name}",
            "items": [{
                "combined_question": ""
            }]
        }

async def generate_combined_for_diagnosis(diagnosis_title: str, sections: list[dict]) -> dict:
    """
    Pair sections under the same diagnosis and generate one combined question per pair.
    Pairs are formed sequentially: (0,1), (2,3), ...
    """
    pairs = []
    for i in range(0, len(sections), 2):
        if i + 1 < len(sections):
            pairs.append((sections[i], sections[i + 1]))
        else:
            if i - 1 >= 0:
                pairs.append((sections[i - 1], sections[i]))

    combined_results = []
    for a, b in pairs:
        res = await generate_combined_questions_for_sections(diagnosis_title, a, b)
        combined_results.append(res)

    return {
        "title": diagnosis_title,
        "combined_items": combined_results
    }

def _ensure_excel_with_headers(xlsx_path: Path, headers: list[str]):
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    if not xlsx_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(headers)
        wb.save(xlsx_path)

def append_rows_to_excel(xlsx_path: Path, rows: list[list]):
    headers = ["file", "title", "section", "question", "question_variant"]
    _ensure_excel_with_headers(xlsx_path, headers)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(xlsx_path)

def rows_from_individual_result(source_file: str, diagnosis_result: dict) -> list[list]:
    rows = []
    title = diagnosis_result.get("title", "")
    for section in diagnosis_result.get("sections", []):
        section_name = section.get("section", "")
        items = section.get("items", [])
        if not items or not isinstance(items[0], dict):
            continue
        entry = items[0]
        # basic_questions (3 rows, variant = basic_question)
        for q in entry.get("basic_questions", []) or []:
            if isinstance(q, str) and q.strip():
                rows.append([source_file, title, section_name, q.strip(), "basic_question"])
        # validation_question (1 row, variant = validation_question)
        vq = entry.get("validation_question", "")
        if isinstance(vq, str) and vq.strip():
            rows.append([source_file, title, section_name, vq.strip(), "validation_question"])
        # deepdive_question (1 row, variant = deepdive_question)
        dq = entry.get("deepdive_question", "")
        if isinstance(dq, str) and dq.strip():
            rows.append([source_file, title, section_name, dq.strip(), "deepdive_question"])
    return rows

def rows_from_combined_result(source_file: str, combined_result: dict) -> list[list]:
    rows = []
    title = combined_result.get("title", "")
    for item in combined_result.get("combined_items", []):
        section = item.get("combined_section", "")
        items = item.get("items", [])
        if isinstance(items, list) and items and isinstance(items[0], dict):
            cq = items[0].get("combined_question", "")
            if isinstance(cq, str) and cq.strip():
                rows.append([source_file, title, section, cq.strip(), "combined_question"])
    return rows

async def process_combined_output_for_file(in_path: Path, out_path: Path):
    """
    For a single input file, write combined outputs (single question per pair) into output_files/combined_output.
    """
    log_info(f"Processing combined output for: {in_path.name}")
    try:
        with in_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        log_error(f"Error reading file {in_path}", e)
        return

    COMBINED_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = COMBINED_OUTPUT_DIR / in_path.name
    results = {"file": in_path.name, "diagnoses": []}
    if out_path.exists():
        try:
            with out_path.open("r", encoding="utf-8") as f:
                existing = json.load(f)
                if isinstance(existing, dict) and existing.get("file") == in_path.name:
                    results["diagnoses"] = existing.get("diagnoses", [])
        except Exception as e:
            log_error(f"Could not read existing combined output {out_path}", e)

    diagnoses = data.get("diagnosis", [])
    log_info(f"Found {len(diagnoses)} diagnoses in file for combined output")

    for diag in diagnoses:
        title = diag.get("title", "")
        content = diag.get("content", "")
        log_info(f"Combined generation start: {title}")

        sections = extract_sections(content)
        if not sections or len(sections) < 1:
            log_info(f"No sections found for combined generation: {title}")
            continue

        combined_result = await generate_combined_for_diagnosis(title, sections)
        results["diagnoses"].append(combined_result)

        try:
            with out_path.open("w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            log_info(f"Appended combined diagnosis '{title}' to: {out_path}")
        except Exception as e:
            log_error("Error saving combined partial results", e)

        # Append rows to Combined Excel immediately
        try:
            combined_rows = rows_from_combined_result(in_path.name, combined_result)
            if combined_rows:
                append_rows_to_excel(COMBINED_XLSX, combined_rows)
                log_info(f"Excel appended (combined): {len(combined_rows)} rows")
        except Exception as e:
            log_error("Error appending to combined Excel", e)

    log_info(f"Combined processing completed: {len(results['diagnoses'])} diagnoses")

async def process_input_file(in_path: Path, out_path: Path):
    log_info(f"Processing file: {in_path.name}")

    # Load input JSON
    try:
        with in_path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        log_error(f"Error reading file {in_path}", e)
        return

    # Prepare output accumulator: preserve existing output and append
    INDIVIDUAL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = INDIVIDUAL_OUTPUT_DIR / in_path.name
    results = {"file": in_path.name, "diagnoses": []}
    if out_path.exists():
        try:
            with out_path.open("r", encoding="utf-8") as f:
                existing = json.load(f)
                if isinstance(existing, dict) and existing.get("file") == in_path.name:
                    results["diagnoses"] = existing.get("diagnoses", [])
        except Exception as e:
            log_error(f"Could not read existing output {out_path}", e)

    # Process each diagnosis
    diagnoses = data.get("diagnosis", [])
    log_info(f"Found {len(diagnoses)} diagnoses in file")

    for diag in diagnoses:
        title = diag.get("title", "")
        content = diag.get("content", "")

        log_info(f"Processing diagnosis: {title}")

        sections = extract_sections(content)
        if not sections:
            log_info(f"No sections found for {title}")
            continue

        log_info(f"Found {len(sections)} sections for {title}")
        for sec in sections:
            log_info(f"  Section: {sec['section_name']} (len={len(sec['section_text'])})")

        diagnosis_result = await generate_questions_for_diagnosis(title, sections)
        # Always append and write immediately
        results["diagnoses"].append(diagnosis_result)
        try:
            with out_path.open("w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            log_info(f"Appended diagnosis '{title}' to: {out_path}")
        except Exception as e:
            log_error("Error saving partial results", e)

        # Append rows to Individual Excel immediately
        try:
            indiv_rows = rows_from_individual_result(in_path.name, diagnosis_result)
            if indiv_rows:
                append_rows_to_excel(INDIVIDUAL_XLSX, indiv_rows)
                log_info(f"Excel appended (individual): {len(indiv_rows)} rows")
        except Exception as e:
            log_error("Error appending to individual Excel", e)

    log_info(f"Processed {len(results['diagnoses'])} diagnoses successfully")

async def main():
    """Process all JSON files in the input directory"""
    input_files = list(CAREGUIDES_INPUT.glob("*.json"))
    log_info(f"Found {len(input_files)} JSON files to process")
    
    for in_file in input_files:
        # Individual per-section outputs -> output_files/individual_output
        await process_input_file(in_file, INDIVIDUAL_OUTPUT_DIR / in_file.name)

        # Combined outputs (single question per pair) -> output_files/combined_output
        await process_combined_output_for_file(in_file, COMBINED_OUTPUT_DIR / in_file.name)
    
    log_info("Processing complete!")

if __name__ == "__main__":
    asyncio.run(main())




